import tkinter as tk
import tkinter.ttk as ttk
import pandas as pd
import openpyxl
import datetime

win = tk.Tk()

win.title ("Photo_manager_2022")
win.geometry('450x600')

# 1 ФУНКЦИИ
def data_del():

    wb_list.delete_rows(b)
    wb.save(filename='D:\kursovaja\photo_2022\photo_data_new.xlsx')


def otrisovka_suz_dan(city_old,name_old,tel_old,price_old,time_old,start_old):
    city_old_1 = tk.Label(win, text=city_old)
    city_old_1.grid(row=4, column=1)
    name_old_1 = tk.Label(win, text = name_old)
    name_old_1.grid(row=5, column=1)
    tel_old_1 = tk.Label(win, text=tel_old)
    tel_old_1.grid(row=6, column=1)
    price_old_1 = tk.Label(win, text=price_old)
    price_old_1.grid(row=7, column=1)
    time_old_1 = tk.Label(win, text=time_old)
    time_old_1.grid(row=8, column=1)
    start_old_1 = tk.Label(win, text=start_old)
    start_old_1.grid(row=9, column=1)
    city_lab.grid(row=4, column=0, padx=10, pady=10)
    name_lab.grid(row=5, column=0, padx=10, pady=10)
    tel_lab.grid(row=6, column=0, padx=10, pady=10)
    price_lab.grid(row=7, column=0, padx=10, pady=10)
    time_lab.grid(row=8, column=0, padx=10, pady=10)
    start_lab.grid(row=9, column=0, padx=10, pady=10)
    hint_3.grid(row=10, column=0, padx=10, pady=10)
    btn_2_yes.grid(row=11, column=0, padx=10, pady=10)
    btn_2_NO.grid(row=11, column=1, padx=10, pady=10)
    btn_2_del.grid(row=11, column=2, padx=10, pady=10)

def otrisovka_dan_zap():
    city_lab.grid(row=6, column=0, padx=10, pady=10)
    name_lab.grid(row=7, column=0, padx=10, pady=10)
    tel_lab.grid(row=8, column=0, padx=10, pady=10)
    price_lab.grid(row=9, column=0, padx=10, pady=10)
    time_lab.grid(row=10, column=0, padx=10, pady=10)
    start_lab.grid(row=11, column=0, padx=10, pady=10)
    # Полей ввода
    city_ent.grid(row=6, column=1, padx=10, pady=10)
    name_ent.grid(row=7, column=1, padx=10, pady=10)
    tel_ent.grid(row=8, column=1, padx=10, pady=10)
    price_ent.grid(row=9, column=1, padx=10, pady=10)
    time_ent.grid(row=10, column=1, padx=10, pady=10)
    start_ent.grid(row=11, column=1, padx=10, pady=10)
    # Отрисовка кнопки записи
    btn_zapis.grid(row=12, column=0, columnspan=3, stick='we', padx=10, pady=10)
def proverka():
    day = combo_days.get()
    mouth = combo_mounth.get()
    year = combo_years.get()
    out_2 = day,mouth,year
    out_2_1 = "Дата:",out_2,"уже ЗАБРОНИРОВАННА"
    out_2_2 = "Дата:",out_2,"СВОБОДНА! Хотите записать её ?"
    global wb
    wb = openpyxl.reader.excel.load_workbook(filename='D:\kursovaja\photo_2022\photo_data_new.xlsx')
    global wb_list
    wb_list = wb[year]
    # a - поиск количества строк
    a = wb_list.max_row
    mounth_num = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)
    mouth_sl = dict(zip(mounth, mounth_num))
    # искомая датаsearcher
    sercher = datetime.datetime(int(year), mouth_sl[mouth], int(day))
    for i in range(2, a + 1):

        # print(wb_list['B'+str(i)].value, "дата:",wb_list['A'+str(i)].value )
        val = wb_list.cell(row=i, column=1).value

        if val == sercher:
            global b
            b = i
            print("Есть совпадение, по дате :", wb_list['A' + str(i)].value)
            city_old = wb_list['B'+str(i)].value
            name_old = wb_list['C'+str(i)].value
            tel_old  = wb_list['D'+str(i)].value
            price_old = wb_list['E'+str(i)].value
            time_old = wb_list['F'+str(i)].value
            start_old = wb_list['G'+str(i)].value

            # Отрисовка кнопок перезаписи
            d_2 = tk.Label(win, text= out_2_1)
            d_2.grid(row=3, column=0, padx=10, pady=10)
            otrisovka_suz_dan(city_old,name_old,tel_old,price_old,time_old,start_old)

            break
    else:
        d_3 = tk.Label(win, text=out_2_2)
        d_3.grid(row=3, column=0, padx=10, pady=10)
        hint_2.grid(row=3, column=0, padx=10, pady=10)
        btn_3_yes.grid(row = 5 , column=0, padx=10, pady=10)
        btn_3_NO.grid(row = 5 , column=1, padx=10, pady=10)
#2 Кнопки/Лэйбы

# 2.1  Выпадающая дата
days = list(range(1, 32))
combo_days = ttk.Combobox(win,values=days, width = 4)
combo_days.current(0)
combo_days.grid(row = 1, column=0, padx=10, pady=10,stick = 'we')

#2.2  Выпадающий список с месяцами
mounth= ("ЯНВАРЬ","ФЕВРАЛЬ","МАРТ","АПРЕЛЬ","МАЙ","ИЮНЬ","ИЮЛЬ","АВГУСТ","СЕНТЯБРЬ","ОКТЯБРЬ","НОЯБРЬ","ДЕКАБРЬ")
combo_mounth = ttk.Combobox(win, values=mounth, width = 20)
combo_mounth.current(0)
combo_mounth.grid(row = 1, column=1, padx=10, pady=10,stick = 'w')

# 2.3 Выпадающий год
years = list(range(2022,2025))
combo_years = ttk.Combobox(win,values=years,  width = 10)
combo_years.current(0)
combo_years.grid(row = 1, column=2, padx=10, pady=10,stick = 'we')

# 2. 4 Кнопка забора даты / "Проверка даты"
btn_vvod = tk.Button(win,text= "Проверить дату",command= proverka,)
btn_vvod.grid(row = 2, column=0,columnspan=3,stick = 'we',padx=10, pady=10)

# 2.5. Верхняя подсказка
hint_1 = tk.Label(win, text= "Выберите желаемую дату:")
hint_1.grid(row = 0 , column=0, padx=10, pady=10)

# 2.6 Подсказак/ Записать дату ?
hint_2 = tk.Label(win, text= "Вы хотите записать дату?")

# 2. 7 Подсказка / Вы хотите перезаписать дату ?
hint_3 = tk.Label(win, text= "Вы хотите ПЕРЕЗАПИСАТЬ дату?")


#2.8 Кнопки выбора
btn_2_yes = tk.Button(win,text= "ДА",command= otrisovka_suz_dan,)
btn_2_NO = tk.Button(win,text= "Нет",command= proverka,)
btn_2_del = tk.Button(win,text= "Удалить дату",command= data_del,)
btn_3_zapis = tk.Button(win,text= "Записать дату",command= proverka,)
btn_3_yes = tk.Button(win,text= "ДА",command= otrisovka_dan_zap,)
btn_3_NO = tk.Button(win,text= "Нет",command= proverka,)

# 2.9 Кнопки заполнения (город,начало)
# Лэйбы
city_lab = tk.Label(win, text= "Город")
name_lab = tk.Label(win, text= "Имя заказчика:")
tel_lab = tk.Label(win, text= "Контактный телефон с кодом:")
price_lab = tk.Label(win, text= "Стоимость")
time_lab = tk.Label(win, text= "Продолжительность:")
start_lab = tk.Label(win, text= "Начало мероприятия:")
# Строки ввода данных
city_ent = tk.Entry(win)
name_ent = tk.Entry(win)
tel_ent = tk.Entry(win)
price_ent = tk.Entry(win)
time_ent = tk.Entry(win)
start_ent = tk.Entry(win)

# 2.10 Кнопка записи
btn_zapis = tk.Button(win,text= "Записать новую дату ",command= proverka,)


win.mainloop()